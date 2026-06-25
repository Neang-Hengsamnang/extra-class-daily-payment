import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, current_app
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.exc import IntegrityError
from extensions import db
from models import Student, Course, Payment, PaymentCourse, User
from datetime import date, datetime, timedelta
import qrcode
from io import BytesIO
from functools import wraps

admin_bp = Blueprint('admin', __name__)

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.role != 'admin':
            flash('Access denied.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated

@admin_bp.route('/')
@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    return render_template('admin/dashboard.html')

# ---- Student Management ----
@admin_bp.route('/students')
@admin_required
def students():
    all_students = Student.query.order_by(Student.grade_level, Student.id).all()
    courses = Course.query.all()
    return render_template('admin/students.html', students=all_students, courses=courses)

@admin_bp.route('/student/add', methods=['POST'])
@admin_required
def add_student():
    full_name = request.form.get('full_name')
    default_course_id = request.form.get('default_course_id')
    if not full_name or not default_course_id:
        flash('ឈ្មោះពេញ និងវគ្គសិក្សាគោលដែលត្រូវការ។', 'danger')
        return redirect(url_for('admin.students'))

    # Generate new student ID
    last = Student.query.order_by(Student.id.desc()).first()
    if last:
        num = int(last.id[4:]) + 1
    else:
        num = 1
    student_id = f"Stu-{num:04d}"

    # Generate QR code
    qr = qrcode.make(student_id)
    qr_dir = os.path.join(current_app.static_folder, 'qrcodes')
    
    # Handle case where qrcodes exists as a file instead of directory
    if os.path.isfile(qr_dir):
        os.remove(qr_dir)
    
    os.makedirs(qr_dir, exist_ok=True)
    filename = f"{student_id}.png"
    qr.save(os.path.join(qr_dir, filename))

    gender = request.form.get('gender')
    grade_level = request.form.get('grade_level')
    student = Student(
        id=student_id,
        full_name=full_name,
        gender=gender,
        grade_level=grade_level,
        default_course_id=int(default_course_id),
        qr_code_filename=filename,
        registration_date=date.today()
    )
    db.session.add(student)
    db.session.commit()
    flash(f'សិស្សឈ្មោះ {full_name} បានបន្ថែមដោយ ID {student_id}.', 'success')
    return redirect(url_for('admin.students'))

@admin_bp.route('/student/edit/<string:id>', methods=['POST'])
@admin_required
def edit_student(id):
    student = Student.query.get_or_404(id)
    student.full_name = request.form.get('full_name', student.full_name)
    student.gender = request.form.get('gender', student.gender)
    student.grade_level = request.form.get('grade_level', student.grade_level)
    student.default_course_id = int(request.form.get('default_course_id', student.default_course_id))
    # Fix: Check if checkbox is in form (means it was checked)
    student.is_active = 'is_active' in request.form and request.form.get('is_active') == 'true'
    db.session.commit()
    flash('សិស្សបានធ្វើបច្ចុប្បន្នភាព។', 'success')
    return redirect(url_for('admin.students'))

@admin_bp.route('/student/deactivate/<string:id>')
@admin_required
def deactivate_student(id):
    student = Student.query.get_or_404(id)
    student.is_active = False
    db.session.commit()
    flash(f'សិស្ស {student.full_name} បានដាក់ឱ្យផ្អាកឬឈប់រៀន។', 'info')
    return redirect(url_for('admin.students'))

@admin_bp.route('/student/activate/<string:id>')
@admin_required
def activate_student(id):
    student = Student.query.get_or_404(id)
    student.is_active = True
    db.session.commit()
    flash(f'សិស្ស {student.full_name} បានដាក់ឱ្យនៅរៀនវិញ។', 'success')
    return redirect(url_for('admin.students'))

@admin_bp.route('/student/delete/<string:id>')
@admin_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    student_name = student.full_name
    db.session.delete(student)
    db.session.commit()
    flash(f'សិស្ស {student_name} បានលុបយ៉ាងជោគជ័យ។', 'success')
    return redirect(url_for('admin.students'))

# ---- Course Management ----
@admin_bp.route('/courses')
@admin_required
def courses():
    all_courses = Course.query.all()
    return render_template('admin/courses.html', courses=all_courses)

@admin_bp.route('/course/add', methods=['POST'])
@admin_required
def add_course():
    name = request.form.get('name')
    fee = request.form.get('daily_fee')
    if not name or fee is None:
        flash('ឈ្មោះវគ្គសិក្សា និងតម្លៃ ត្រូវតែមាន!', 'danger')
        return redirect(url_for('admin.courses'))
    course = Course(name=name, daily_fee=float(fee))
    db.session.add(course)
    db.session.commit()
    flash('វគ្គសិក្សាបានបន្ថែម។', 'success')
    return redirect(url_for('admin.courses'))

@admin_bp.route('/course/edit/<int:id>', methods=['POST'])
@admin_required
def edit_course(id):
    course = Course.query.get_or_404(id)
    course.name = request.form.get('name', course.name)
    course.daily_fee = float(request.form.get('daily_fee', course.daily_fee))
    db.session.commit()
    flash('វគ្គសិក្សាបានធ្វើបច្ចុប្បន្នភាព។', 'success')
    return redirect(url_for('admin.courses'))

@admin_bp.route('/course/delete/<int:id>')
@admin_required
def delete_course(id):
    course = Course.query.get_or_404(id)
    course_name = course.name
    db.session.delete(course)
    db.session.commit()
    flash(f'វគ្គសិក្សា {course_name} បានលុប។', 'success')
    return redirect(url_for('admin.courses'))

# ---- Daily Report ----
@admin_bp.route('/daily_report', methods=['GET', 'POST'])
@admin_required
def daily_report():
    selected_date = date.today()
    if request.method == 'POST':
        date_str = request.form.get('report_date')
        if date_str:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()

    active_students = Student.query.filter_by(is_active=True).all()
    payments_today = Payment.query.filter_by(date=selected_date).all()
    paid_dict = {p.student_id: p for p in payments_today}

    present_count = len(payments_today)
    total_active = len(active_students)
    absent_count = total_active - present_count

    revenue_collected = sum(p.total_amount for p in payments_today if p.is_paid)
    tabs_outstanding = sum(p.total_amount for p in payments_today if not p.is_paid)

    report_rows = []
    for s in active_students:
        payment = paid_dict.get(s.id)
        if payment:
            courses_taken = ', '.join([pc.course.name for pc in payment.courses])
            amount = f"៛{payment.total_amount:,.0f} {'(Paid)' if payment.is_paid else '(Tabs)'}"
            status = 'Present'
        else:
            courses_taken = '-'
            amount = '-'
            status = 'Absent'
        report_rows.append({
            'student': s,
            'status': status,
            'courses_taken': courses_taken,
            'amount': amount
        })

    return render_template('admin/daily_report.html',
                           report_rows=report_rows,
                           selected_date=selected_date,
                           total_active=total_active,
                           present_count=present_count,
                           absent_count=absent_count,
                           revenue_collected=revenue_collected,
                           tabs_outstanding=tabs_outstanding)

# ---- Monthly Report ----
@admin_bp.route('/monthly_report', methods=['GET'])
@admin_required
def monthly_report():
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', date.today().month, type=int)

    # All active students at the time of report (simplified)
    active_count = Student.query.filter_by(is_active=True).count()

    # Query all payments in that month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    payments = Payment.query.filter(Payment.date >= start_date, Payment.date < end_date).all()

    # Daily revenue
    daily_revenue = {}
    daily_present = {}
    for p in payments:
        d = p.date
        daily_revenue[d] = daily_revenue.get(d, 0) + p.total_amount
        daily_present[d] = daily_present.get(d, 0) + 1

    # Build complete day series
    import calendar
    days_in_month = calendar.monthrange(year, month)[1]
    labels = [f"{year}-{month:02d}-{d:02d}" for d in range(1, days_in_month+1)]
    revenue_series = []
    present_series = []
    absent_series = []
    for d in range(1, days_in_month+1):
        dt = date(year, month, d)
        rev = daily_revenue.get(dt, 0)
        pres = daily_present.get(dt, 0)
        revenue_series.append(round(rev, 2))
        present_series.append(pres)
        absent_series.append(active_count - pres)

    # Revenue by course
    course_revenue = {}
    for p in payments:
        for pc in p.courses:
            cname = pc.course.name
            course_revenue[cname] = course_revenue.get(cname, 0) + pc.course.daily_fee

    # Paid vs Tabs pie
    total_paid = sum(p.total_amount for p in payments if p.is_paid)
    total_tabs = sum(p.total_amount for p in payments if not p.is_paid)

    # Summary stats
    total_revenue = total_paid + total_tabs
    average_attendance = round(sum(present_series) / days_in_month, 1) if days_in_month else 0

    return render_template('admin/monthly_report.html',
                           year=year, month=month,
                           labels=labels,
                           revenue_series=revenue_series,
                           present_series=present_series,
                           absent_series=absent_series,
                           course_revenue_keys=list(course_revenue.keys()),
                           course_revenue_values=list(course_revenue.values()),
                           total_paid=total_paid,
                           total_tabs=total_tabs,
                           total_revenue=total_revenue,
                           average_attendance=average_attendance)

# ---- User Management (Admin can add staff) ----
@admin_bp.route('/users')
@admin_required
def users():
    all_users = User.query.all()
    return render_template('admin/users.html', users=all_users)

@admin_bp.route('/user/add', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    if not username or not password or role not in ('admin', 'staff'):
        flash('គ្រប់ប្រអប់ទិន្នន័យត្រូវតែបំពេញ!', 'danger')
        return redirect(url_for('admin.users'))
    if User.query.filter_by(username=username).first():
        flash('Username មានគេប្រើរួចហើយ!', 'danger')
        return redirect(url_for('admin.users'))
    user = User(username=username, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash('អ្នកប្រើប្រាស់ត្រូវបានបន្ថែមរួចរាល់ ។', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/user/edit/<int:id>', methods=['POST'])
@admin_required
def edit_user(id):
    user = User.query.get_or_404(id)
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    
    if not username or role not in ('admin', 'staff'):
        flash('គ្រប់ប្រអប់ទិន្នន័យត្រូវតែបំពេញ!', 'danger')
        return redirect(url_for('admin.users'))
    
    # Check if new username already exists (and it's not the same user)
    if username != user.username and User.query.filter_by(username=username).first():
        flash('Username មានគេប្រើរួចហើយ!', 'danger')
        return redirect(url_for('admin.users'))
    
    user.username = username
    user.role = role
    
    # Update password only if provided
    if password:
        user.set_password(password)
    
    db.session.commit()
    flash('ទិន្នន័យអ្នកប្រើប្រាស់ត្រូវបានធ្វើបច្ចុប្បន្នភាពដោយជោគជ័យ ។', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/user/delete/<int:id>')
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    
    # Prevent deleting the current user
    if user.id == current_user.id:
        flash('មិនអាចលុបគណនីផ្ទាល់ខ្លួនឯងបានទេ!', 'danger')
        return redirect(url_for('admin.users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'User {username} deleted.', 'success')
    return redirect(url_for('admin.users'))

@admin_bp.route('/data-management')
@login_required
def data_management():
    """Admin page for importing/exporting student data."""
    if current_user.role != 'admin':
        flash('សិទ្ធចូលប្រើត្រូវបានបដិសេធ. មានសិទ្ធបានតែ Admin ទេ។', 'error')
        return redirect(url_for('staff.scan'))
    return render_template('admin/data_management.html')


@admin_bp.route('/export-students')
@login_required
def export_students():
    """Export all students to an Excel file with columns: id, full_name, gender, grade_level."""
    if current_user.role != 'admin':
        flash('សិទ្ធចូលប្រើត្រូវបានបដិសេធ. មានសិទ្ធបានតែ Admin ទេ។', 'error')
        return redirect(url_for('staff.scan'))

    students = Student.query.order_by(Student.grade_level, Student.full_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Headers matching your Student model fields
    headers = ['id', 'full_name', 'gender', 'grade_level', 'default_course']
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Data rows
    for student in students:
        course_name = student.default_course.name if student.default_course else ''
        ws.append([
            student.id,
            student.full_name,
            student.gender or '',
            student.grade_level or '',
            course_name
        ])

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"students_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_bp.route('/import-students', methods=['POST'])
@login_required
def import_students():
    """Import students from an uploaded Excel file.
       Updates existing records if student id exists, otherwise creates new.
    """
    if current_user.role != 'admin':
        flash('សិទ្ធចូលប្រើត្រូវបានបដិសេធ. មានសិទ្ធបានតែ Admin ទេ។', 'error')
        return redirect(url_for('staff.scan'))

    # Validate file
    if 'file' not in request.files:
        flash('មិនមាន file ណាមួយត្រូវបាន Upload!', 'error')
        return redirect(url_for('admin.data_management'))

    file = request.files['file']
    if file.filename == '':
        flash('មិនមាន file ណាមួយត្រូវបានជ្រើសរើស!', 'error')
        return redirect(url_for('admin.data_management'))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('សូម upload file Excel មួយមក (.xlsx or .xls).', 'error')
        return redirect(url_for('admin.data_management'))

    # Load workbook
    try:
        workbook = load_workbook(file)
        sheet = workbook.active
    except Exception as e:
        flash(f'បញ្ហាក្នុងការអានឯកសារ/Error reading file: {str(e)}', 'error')
        return redirect(url_for('admin.data_management'))

    # Validate headers (case-insensitive)
    expected_headers = ['id', 'full_name', 'gender', 'grade_level']
    first_row = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
    first_row_lower = [h.lower() for h in first_row]
    expected_lower = [h.lower() for h in expected_headers]

    # We'll accept either 'default_course' or 'course' as the column name
    all_possible = ['id', 'full_name', 'gender', 'grade_level', 'default_course', 'course']

    # Check that at least 'id', 'full_name' are present (others optional)
    mandatory = ['id', 'full_name']
    if not all(m in first_row_lower for m in [m.lower() for m in mandatory]):
        flash(f'ទ្រង់ទ្រាយការបញ្ចូលមិនត្រឹមត្រូវ ។ ត្រូវបញ្ចូលយ៉ាងហោចណាស់ "{", ".join(mandatory)}" headers.', 'error')
        flash(f'Found: {first_row}', 'error')
        return redirect(url_for('admin.data_management'))

    # Map column indices (case-insensitive) for all possible fields
    col_indices = {}
    for col_name in all_possible:
        for idx, header in enumerate(first_row):
            if header.lower() == col_name.lower():
                col_indices[col_name] = idx
                break
    # If 'default_course' not found but 'course' is, map it
    if 'default_course' not in col_indices and 'course' in col_indices:
        col_indices['default_course'] = col_indices['course']
    # Remove duplicate key if both exist
    col_indices.pop('course', None)

    students_added = 0
    students_updated = 0
    errors = []

    courses = {c.name: c.id for c in Course.query.all()}
    course_warnings = []   # add this near other accumulators like errors

    # Ensure QR codes directory exists (same logic as add_student)
    qr_dir = os.path.join(current_app.static_folder, 'qrcodes')
    # Handle case where qrcodes exists as a file instead of directory
    if os.path.isfile(qr_dir):
        os.remove(qr_dir)
    os.makedirs(qr_dir, exist_ok=True)

    def generate_qr(student_id):
        """Generate QR code for the given student ID and return filename."""
        filename = f"{student_id}.png"
        qr = qrcode.make(student_id)
        qr.save(os.path.join(qr_dir, filename))
        return filename

    # Iterate rows (skip header)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        # Skip completely empty rows
        if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
            continue

        try:
            # Extract data using column indices (if column missing, default to empty)
            student_id = str(row[col_indices['id']].value).strip() if 'id' in col_indices and row[col_indices['id']].value else ''
            full_name = str(row[col_indices['full_name']].value).strip() if 'full_name' in col_indices and row[col_indices['full_name']].value else ''
            gender = str(row[col_indices['gender']].value).strip() if 'gender' in col_indices and row[col_indices['gender']].value else ''
            grade_level = str(row[col_indices['grade_level']].value).strip() if 'grade_level' in col_indices and row[col_indices['grade_level']].value else ''
            course_name = str(row[col_indices['default_course']].value).strip() if 'default_course' in col_indices and row[col_indices['default_course']].value else ''

            # Validate mandatory fields
            if not student_id:
                errors.append(f"Row {row_idx}: Missing Student ID")
                continue
            if not full_name:
                errors.append(f"Row {row_idx}: Missing Full Name")
                continue

            course_id = None
            if course_name:
                course_id = courses.get(course_name)   # returns None if not found
                if course_name and not course_id:
                    course_warnings.append(f"Row {row_idx}: Course '{course_name}' not found.")

            
            # If student_id is missing or empty, generate a new one (like add_student)
            if not student_id:
                last = Student.query.order_by(Student.id.desc()).first()
                if last:
                    num = int(last.id[4:]) + 1
                else:
                    num = 1
                student_id = f"Stu-{num:04d}"
            
            # Find existing or create new
            existing = Student.query.get(student_id)
            if existing:
                existing.full_name = full_name
                existing.gender = gender if gender else None
                existing.grade_level = grade_level if grade_level else None
                existing.default_course_id = course_id
                # Ensure QR code exists (if missing, generate it)
                qr_file = os.path.join(current_app.static_folder, 'qrcodes', f"{student_id}.png")
                if not os.path.exists(qr_file):
                    try:
                        generate_qr(student_id)
                    except:
                        pass  # or log warning
                students_updated += 1
            else:
                # Generate QR code (same as add_student)
                try:
                    qr_filename = generate_qr(student_id)
                except Exception as e:
                    errors.append(f"Row {row_idx}: QR generation failed: {str(e)}")
                    qr_filename = None   # still add student but without QR

                new_student = Student(
                    id=student_id,
                    full_name=full_name,
                    gender=gender if gender else None,
                    grade_level=grade_level if grade_level else None,
                    default_course_id=course_id,
                    qr_code_filename=qr_filename,   # set the field
                    registration_date=date.today()
                )
                db.session.add(new_student)
                students_added += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            continue

    # Commit changes
    try:
        db.session.commit()
        if errors:
            flash(f'⚠️ {len(errors)} error(s) encountered. First 5 shown:', 'warning')
            for err in errors[:5]:
                flash(f'• {err}', 'error')
            if len(errors) > 5:
                flash(f'... and {len(errors)-5} more', 'error')
        if course_warnings:
            flash(f'⚠️ {len(course_warnings)} course not found warning(s):', 'warning')
            for warn in course_warnings[:5]:
                flash(f'• {warn}', 'warning')
            if len(course_warnings) > 5:
                flash(f'... and {len(course_warnings)-5} more', 'warning')
        flash(f'✅ ការនាំចូលបានបញ្ចប់រួចរាល់! ដែលបានបន្ថែម: {students_added}, និងបានធ្វើបច្ចុប្បន្នភាព: {students_updated}', 'success')
    except IntegrityError as e:
        db.session.rollback()
        flash(f'❌ បញ្ហាជាមួយមូលដ្ឋានទិន្នន័យ (Database): {str(e)}', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ បញ្ហាមិនបានរំពឹងទុកមុន: {str(e)}', 'error')

    return redirect(url_for('admin.data_management'))

@admin_bp.route('/print-qr-cards')
@login_required
@admin_required
def print_qr_cards():
    """Render a print-friendly page with QR code cards for selected students."""
    # Get student IDs from query string: ?ids=Stu-0001,Stu-0002
    ids_param = request.args.get('ids', '')
    if ids_param:
        # Split by comma and strip whitespace
        student_ids = [id.strip() for id in ids_param.split(',') if id.strip()]
        students = Student.query.filter(Student.id.in_(student_ids)).order_by(Student.grade_level, Student.full_name).all()
    else:
        # Print all active students (or all if you prefer)
        students = Student.query.filter_by(is_active=True).order_by(Student.grade_level, Student.full_name).all()

    # If no students, flash and redirect back
    if not students:
        flash('No students selected or found to print.', 'warning')
        return redirect(url_for('admin.students'))

    return render_template('admin/print_qr_cards.html', students=students)